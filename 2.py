import os
import sys

import numpy as np
import plotly.offline as offline
import pandas as pd
import plotly.express as px
from PyQt6.QtCore import *
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from openpyxl.reader.excel import load_workbook
from forecast import main_import_file


class PlotDialog(QDialog):
    def __init__(self, plot_html_file):
        super(PlotDialog, self).__init__()
        self.setWindowTitle('График')
        self.resize(800, 600)

        # Создаем QWebEngineView для отображения графика
        self.webview = QWebEngineView()
        self.webview.setUrl(QUrl.fromLocalFile(os.path.abspath(plot_html_file)))

        # Создаем вертикальный лейаут
        layout = QVBoxLayout()
        layout.addWidget(self.webview)

        # Устанавливаем лейаут для диалогового окна
        self.setLayout(layout)


class FileSelectionApp(QMainWindow):
    def __init__(self):
        super(FileSelectionApp, self).__init__(None)
        self.selected_file = None  # Объявляем selected_file как атрибут экземпляра класса
        self.selected_sheet = None  # Объявляем selected_sheet
        self.setWindowTitle('File Selection App')
        self.resize(1200, 800)
        # Создаем вертикальный лейаут для размещения виджетов
        layout = QVBoxLayout()

        # Создаем QLabel для отображения выбранного файла
        self.file_label = QLabel('Выберите файл:')
        layout.addWidget(self.file_label)

        # Создаем кнопку для создания прогноза
        self.create_prediction_on_chosen_file = QPushButton('Создать прогноз')
        self.create_prediction_on_chosen_file.setEnabled(False)
        # Используем lambda-функцию, чтобы передать параметр
        self.create_prediction_on_chosen_file.clicked.connect(self.create_prediction)
        layout.addWidget(self.create_prediction_on_chosen_file)

        # Создаем кнопку для выбора листа
        self.select_sheet_button = QPushButton('Выбрать лист')
        self.select_sheet_button.setEnabled(False)
        self.select_sheet_button.clicked.connect(self.choose_excel_sheet)
        layout.addWidget(self.select_sheet_button)

        # Создаем кнопку для создания графика
        self.create_prediction_graph = QPushButton('Создать график')
        # Используем lambda-функцию, чтобы передать параметр
        self.create_prediction_graph.clicked.connect(self.show_plot)
        layout.addWidget(self.create_prediction_graph)

        # Создаем QLabel для отображения выбранного листа
        self.sheet_label = QLabel('Выбранный лист:')
        layout.addWidget(self.sheet_label)

        # Создаем QTableWidget для отображения данных
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        # Создаем виджет и устанавливаем наш лейаут
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        # Добавляем меню в верхнюю панель
        menu_bar = self.menuBar()

        # Создаем QAction "File" с выпадающим меню
        file_menu_action = QAction("File", self)
        file_menu = QMenu(self)

        # Добавляем в меню кнопку "Выбрать файл"
        select_file_action = QAction("Выбрать файл", self)
        select_file_action.triggered.connect(self.open_file_and_choose_sheet)
        file_menu.addAction(select_file_action)

        # Создаем QWebEngineView для отображения графика
        self.webview = QWebEngineView()
        layout.addWidget(self.webview)

        file_menu_action.setMenu(file_menu)
        menu_bar.addAction(file_menu_action)

    def create_prediction(self):
        if self.selected_file and self.selected_sheet:
            main_import_file(self.selected_file, self.selected_sheet)

    def open_file_and_choose_sheet(self):
        try:
            # Открываем диалог выбора файла и фильтра для Excel
            file_dialog = QFileDialog(self)
            file_dialog.setNameFilter("Excel Files (*.xlsx)")
            selected_file, _ = file_dialog.getOpenFileName(self, 'Выберите файл', '', 'Excel Files (*.xlsx)')

            # Обновляем QLabel с выбранным файлом
            self.file_label.setText(f'Выбранный файл: {selected_file}')

            # Выбираем лист Excel, если это файл Excel
            if selected_file.endswith('.xlsx'):
                self.selected_file = selected_file
                self.select_sheet_button.setEnabled(True)
                self.create_prediction_on_chosen_file.setEnabled(True)
                self.choose_excel_sheet()

        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")

    def choose_excel_sheet(self):
        try:
            workbook = load_workbook(filename=self.selected_file, read_only=True)
            sheet_names = workbook.sheetnames

            # Используем QInputDialog для выбора листа
            input_dialog = QInputDialog(self)
            input_dialog.setWindowTitle('Выбор листа')
            input_dialog.setComboBoxItems(sheet_names)
            input_dialog.setComboBoxEditable(False)

            # Устанавливаем фиксированный размер
            input_dialog.setFixedSize(400, 200)

            # Отображаем диалоговое окно
            ok_pressed = input_dialog.exec()

            if ok_pressed:
                # Очищаем таблицу перед отображением новых данных
                self.table_widget.clear()
                self.selected_sheet = input_dialog.textValue()
                # Загружаем данные из выбранного листа
                data = pd.read_excel(self.selected_file, self.selected_sheet, engine='openpyxl')
                self.display_data_in_table(data)

                # Обновляем QLabel с выбранным листом
                self.sheet_label.setText(f'Выбранный лист: {self.selected_sheet}')

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def display_data_in_table(self, data):
        try:
            # Устанавливаем количество строк и столбцов
            self.table_widget.setRowCount(data.shape[0])
            self.table_widget.setColumnCount(data.shape[1])

            # Устанавливаем заголовки столбцов
            self.table_widget.setHorizontalHeaderLabels(data.columns)

            # Заполняем ячейки данными
            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    item = QTableWidgetItem(str(data.iloc[i, j]))
                    self.table_widget.setItem(i, j, item)

            # Разрешаем растягивание столбцов
            self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        except Exception as e:
            print(f"Ошибка при отображении данных в таблице: {e}")

    def show_plot(self):
        try:
            if self.selected_file and self.selected_sheet:
                # Здесь вы можете вставить свой код для создания графика с использованием Plotly
                # Ниже приведен пример для демонстрации
                x = np.arange(1000)
                y = x ** 2

                fig = px.scatter(x=x, y=y, labels={'x': 'X Label', 'y': 'Y Label'}, title='Пример графика')

                # we create html code of the figure
                html = '<html><body>'
                html += offline.plot(fig, output_type='div', include_plotlyjs='cdn')
                html += '</body></html>'

                # we create an instance of QWebEngineView and set the html code
                plot_widget = QWebEngineView()
                plot_widget.setHtml(html)

                # set the QWebEngineView instance as main widget
                self.setCentralWidget(plot_widget)

                # Отображаем график в QWebEngineView
                # self.webview.setUrl(QUrl.fromLocalFile(os.path.abspath(plotly_html)))
        except Exception as e:
            print(f"Ошибка при отображении графика: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FileSelectionApp()
    window.show()
    sys.exit(app.exec())
