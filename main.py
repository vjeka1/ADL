import subprocess
import sys
import plotly.offline as offline
import pandas as pd
import plotly.express as px
import logging
import utilities as util
from PyQt6.QtCore import *
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from openpyxl.reader.excel import load_workbook



class FileSelectionApp(QMainWindow):
    """Main window of this program"""

    def __init__(self):
        super(FileSelectionApp, self).__init__(None)
        self.selected_file = None  # Объявляем selected_file как атрибут экземпляра класса
        self.selected_sheet = None  # Объявляем selected_sheet
        self.setWindowTitle('ADL Model')
        self.resize(1200, 800)
        layout = QVBoxLayout()
        self.file_label = QLabel('Выберите файл:')
        layout.addWidget(self.file_label)
        self.select_file_button = QPushButton('Выбрать файл')
        self.select_file_button.clicked.connect(self.open_file_and_choose_sheet)
        layout.addWidget(self.select_file_button)
        self.select_sheet_button = QPushButton('Выбрать лист')
        self.select_sheet_button.setEnabled(False)
        self.select_sheet_button.clicked.connect(self.choose_excel_sheet)
        layout.addWidget(self.select_sheet_button)
        self.create_model_on_chosen_file = QPushButton('Создать модель для прогнозирования')
        self.create_model_on_chosen_file.setEnabled(False)
        self.create_model_on_chosen_file.clicked.connect(self.create_short_term_prediction_model)
        layout.addWidget(self.create_model_on_chosen_file)
        self.create_prediction_on_chosen_file = QPushButton('Создать прогноз')
        self.create_prediction_on_chosen_file.setEnabled(False)
        self.create_prediction_graph = QPushButton('Создать график')
        self.create_prediction_graph.setEnabled(False)
        self.create_prediction_graph.clicked.connect(self.choose_column_name_for_plot)
        layout.addWidget(self.create_prediction_graph)
        self.sheet_label = QLabel('Выбранный лист:')
        layout.addWidget(self.sheet_label)
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)
        menu_bar = self.menuBar()
        file_menu_action = QAction("File", self)
        file_menu = QMenu(self)
        select_file_action = QAction("Выбрать файл", self)
        select_file_action.triggered.connect(self.open_file_and_choose_sheet)
        file_menu.addAction(select_file_action)
        self.webview = QWebEngineView()
        layout.addWidget(self.webview)
        file_menu_action.setMenu(file_menu)
        menu_bar.addAction(file_menu_action)

    def open_file_and_choose_sheet(self):
        try:
            file_dialog = QFileDialog(self)
            file_dialog.setNameFilter("Excel Files (*.xlsx)")
            selected_file, _ = file_dialog.getOpenFileName(self, 'Выберите файл', '', 'Excel Files (*.xlsx)')
            self.file_label.setText(f'Выбранный файл: {selected_file}')
            if selected_file.endswith('.xlsx'):
                self.selected_file = selected_file
                self.select_sheet_button.setEnabled(True)
                self.create_model_on_chosen_file.setEnabled(True)
                self.create_prediction_graph.setEnabled(True)
                self.choose_excel_sheet()
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")

    def choose_excel_sheet(self):
        try:
            workbook = load_workbook(filename=self.selected_file, read_only=True)
            sheet_names = workbook.sheetnames
            input_dialog = QInputDialog(self)
            input_dialog.setLabelText('Выберите лист:')
            input_dialog.setWindowTitle('Выбор листа')
            input_dialog.setComboBoxItems(sheet_names)
            input_dialog.setComboBoxEditable(False)
            input_dialog.setFixedSize(400, 200)
            ok_pressed = input_dialog.exec()
            if ok_pressed:
                self.table_widget.clear()
                self.selected_sheet = input_dialog.textValue()
                data = pd.read_excel(self.selected_file, self.selected_sheet, engine='openpyxl')
                self.display_data_in_table(data)
                self.sheet_label.setText(f'Выбранный лист: {self.selected_sheet}')
        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def display_data_in_table(self, data):
        try:
            self.table_widget.setRowCount(data.shape[0])
            self.table_widget.setColumnCount(data.shape[1])
            self.table_widget.setHorizontalHeaderLabels(data.columns)
            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    item = QTableWidgetItem(str(data.iloc[i, j]))
                    self.table_widget.setItem(i, j, item)
        except Exception as e:
            print(f"Ошибка при отображении данных в таблице: {e}")

    def choose_column_name_for_plot(self):
        try:
            if self.selected_file and self.selected_sheet:
                column_names = [self.table_widget.horizontalHeaderItem(j).text() for j in
                                range(self.table_widget.columnCount())]
                data_selection_dialog = DataSelectionDialogPlot(column_names=column_names,
                                                                selected_file=self.selected_file,
                                                                selected_sheet=self.selected_sheet)
                data_selection_dialog.exec()
        except Exception as e:
            print(f"Ошибка при запуске диалогового окна выбора осей графика: {e}")

    def create_short_term_prediction_model(self):
        try:
            if self.selected_file and self.selected_sheet:
                column_names = [self.table_widget.horizontalHeaderItem(j).text() for j in
                                range(self.table_widget.columnCount())]
                data_selection_dialog = ModelWindow(column_names=column_names, selected_file=self.selected_file,
                                                       selected_sheet=self.selected_sheet)
                data_selection_dialog.exec()
        except Exception as e:
            print(f"Ошибка при запуске диалогового окна выбора настройки создания прогноза: {e}")


class ModelWindow(QDialog):
    """In this window created predicts"""

    def __init__(self, column_names, selected_file, selected_sheet):
        super().__init__()
        self.count_factors = None
        self.column_names = column_names
        self.selected_file = selected_file
        self.selected_sheet = selected_sheet
        self.setWindowTitle('Выбор данных для создания модели')
        self.adjustSize()
        layout = QVBoxLayout()
        self.label_1 = QLabel('Выберите столбец, для которого необходимо составить прогнозную модель:')
        self.menu_data_prediction_name_column = QComboBox()
        self.menu_data_prediction_name_column.addItems(self.column_names)
        self.label_2 = QLabel('Выберите столбец данных с временными метками:')
        self.menu_time_label_name_column = QComboBox()
        self.menu_time_label_name_column.addItems(self.column_names)
        self.label_3 = QLabel('Выберите до 3-ёх факторов прогнозной модели:')
        self.label_4 = QLabel("Выберите размер тестовой выборки:")
        self.slider_box = QSlider()
        self.slider_box.setMinimum(0)
        self.slider_box.setMaximum(100)
        self.slider_box.setOrientation(Qt.Orientation.Horizontal)
        self.slider_value = 66
        self.slider_box.setValue(self.slider_value)
        self.percent_label = QLabel(f"Обучающая выборка: {self.slider_value}% Тестовая выборка: {100-self.slider_value}%")
        self.slider_box.valueChanged.connect(self.slider_value_changed)
        self.label_5 = QLabel('Количество дней прогнозирования:')
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(1)
        self.table_widget.setRowCount(len(column_names)-1)
        self.table_widget.setHorizontalHeaderLabels(['Выбрать столбец'])
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table_widget.itemChanged.connect(self.handle_item_changed)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.menu_time_label_name_column.currentIndexChanged.connect(self.update_table_widget)
        self.update_table_widget()
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(self.label_1)
        layout.addWidget(self.menu_data_prediction_name_column)
        layout.addWidget(self.label_2)
        layout.addWidget(self.menu_time_label_name_column)
        layout.addWidget(self.label_3)
        layout.addWidget(self.table_widget)
        layout.addWidget(self.label_4)
        layout.addWidget(self.slider_box)
        layout.addWidget(self.percent_label)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def update_table_widget(self):
        selected_time_label = self.menu_time_label_name_column.currentText()
        self.list_columns_name = list(self.column_names)
        self.list_columns_name.remove(selected_time_label)
        self.table_widget.setRowCount(len(self.list_columns_name))
        for i, column_name in enumerate(self.list_columns_name):
            item = QTableWidgetItem(column_name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.table_widget.setItem(i, 0, item)

    def slider_value_changed(self, value):
        self.percent_label.setText(f"Обучающая выборка: {value}% Тестовая выборка: {100 - value}%")

    def handle_item_changed(self):
        checked_count = (1 for row in range(self.table_widget.rowCount()) if
                         self.table_widget.item(row, 0).checkState() == Qt.CheckState.Checked)
        try:
            checked_count_sum = sum(checked_count)
            if checked_count_sum > 2:
                for row in range(self.table_widget.rowCount()):
                    item = self.table_widget.item(row, 0)
                    item.setCheckState(Qt.CheckState.Unchecked)
                error_message = ErrorMessageBox('Выбранное количество факторов превышает 2!', self)
                error_message.exec()
        except Exception as e:
            print(f"Exception occurred: {e}")

    def accept(self):
        try:
            chosen_column_for_predict = self.menu_data_prediction_name_column.currentText()
            chosen_column_with_time_label = self.menu_time_label_name_column.currentText()
            if chosen_column_with_time_label == chosen_column_for_predict:
                raise ValueError('Выбранные столбцы совпадают!')
            selected_columns = []
            for row in range(self.table_widget.rowCount()):
                item = self.table_widget.item(row, 0)
                if item is not None and item.checkState() == Qt.CheckState.Checked:
                    selected_columns.append(item.text())
            if not selected_columns:
                raise ValueError('Не выбраны факторы для прогноза!')
            if chosen_column_for_predict in selected_columns:
                create_lag_for_chosen_column_for_predict = True
            else:
                create_lag_for_chosen_column_for_predict = False
                selected_columns.append(chosen_column_for_predict)
            prepared_data = util.data_preparation(
                file=self.selected_file,
                sheet=self.selected_sheet,
                name_column_time=chosen_column_with_time_label,
                name_column_for_predict=chosen_column_for_predict,
                name_column_factors=selected_columns
            )
            data = util.create_lags(
                prepared_data,
                selected_columns,
                1,
                create_lag_for_chosen_column_for_predict,
                chosen_column_for_predict
            )
            data_learn, data_test = util.separation_data(data, self.slider_box.value())
            model = util.create_model(data_learn, chosen_column_for_predict, selected_columns)
            result_data = util.learn_on_params(data, model.params, self.slider_box.value(), chosen_column_for_predict)
            util.calculate_mape(result_data, chosen_column_for_predict)
            result_write_file = util.write_to_excel(
                result_data,
                output_file=chosen_column_for_predict,
                sheet=chosen_column_for_predict
            )
            if result_write_file['Result']:
                question_box = QuestionMessageBox("Открыть созданный файл?", self)
                should_open_file = question_box.exec_and_get_result()
                if should_open_file:
                    subprocess.run(["start", "excel", result_write_file['Path']], shell=True)
                else:
                    pass
            else:
                error_message = ErrorMessageBox('Ошибка записи файла', self)
                error_message.exec()
            super().accept()
        except ValueError as ve:
            error_message = ErrorMessageBox(str(ve), self)
            error_message.exec()

        except Exception as e:
            print(f"Произошла ошибка: {e}")






class DataSelectionDialogPlot(QDialog):
    """In this window is happening choosing data for creating predict model"""

    def __init__(self, column_names, selected_file, selected_sheet):
        super().__init__()
        self.selected_file = selected_file
        self.selected_sheet = selected_sheet
        self.setWindowTitle('Выбор данных для графика')
        self.column_names = column_names
        layout = QVBoxLayout()
        self.label_x = QLabel('Выберите данные для оси X:')
        self.menu_x = QComboBox()
        self.menu_x.addItems(column_names)
        self.menu_x.currentIndexChanged.connect(self.update_table_widget)
        layout.addWidget(self.label_x)
        layout.addWidget(self.menu_x)
        self.label_y = QLabel('Выберите данные для оси Y:')
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(1)
        self.table_widget.setRowCount(len(column_names) - 1)
        self.table_widget.setHorizontalHeaderLabels(['Выбрать столбец'])
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        list_columns_name = list(column_names)
        list_columns_name.remove(self.menu_x.currentText())

        for i, column_name in enumerate(list_columns_name):
            item = QTableWidgetItem(column_name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.table_widget.setItem(i, 0, item)
        layout.addWidget(self.label_y)
        layout.addWidget(self.table_widget)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_selected_data(self):
        selected_x = self.menu_x.currentText()
        selected_y = [self.table_widget.item(i, 0).text() for i in range(self.table_widget.rowCount())
                      if self.table_widget.item(i, 0).checkState() == Qt.CheckState.Checked]
        return selected_x, selected_y

    def update_table_widget(self):
        selected_x = self.menu_x.currentText()
        list_columns_name = list(self.column_names)
        list_columns_name.remove(selected_x)
        self.table_widget.clearContents()
        for i, column_name in enumerate(list_columns_name):
            if column_name != selected_x:
                item = QTableWidgetItem(column_name)
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(Qt.CheckState.Unchecked)
                self.table_widget.setItem(i, 0, item)

    def get_selected_data(self):
        selected_x = self.menu_x.currentText()
        selected_y = [self.table_widget.item(i, 0).text() for i in range(self.table_widget.rowCount())
                      if self.table_widget.item(i, 0).checkState() == Qt.CheckState.Checked]
        return selected_x, selected_y

    def accept(self):
        selected_x, selected_y = self.get_selected_data()
        if selected_x and selected_y:
            data = pd.read_excel(self.selected_file, sheet_name=self.selected_sheet)
            plot_window = PlotWindow(data=data, x_column=selected_x, y_columns=selected_y)
            super().accept()
            plot_window.exec()


class PlotWindow(QDialog):
    """In this window is happening choosing data for creating plot data"""

    def __init__(self, data, x_column, y_columns):
        super().__init__()
        self.setWindowTitle('График')
        self.resize(1200, 800)
        try:
            fig = px.line(data, x=x_column, y=y_columns, title='Название графика')
            html = '<html><body>'
            html += offline.plot(fig, output_type='div', include_plotlyjs='cdn')
            html += '</body></html>'
            plot_widget = QWebEngineView()
            plot_widget.setHtml(html)
            layout = QVBoxLayout()
            layout.addWidget(plot_widget)
            self.setLayout(layout)

        except Exception as e:
            print(f"Ошибка при отображении графика: {e}")


class ErrorMessageBox(QMessageBox):
    def __init__(self, error_message, parent=None):
        super().__init__(parent)
        self.setIcon(QMessageBox.Icon.Critical)
        self.setWindowTitle('Ошибка')
        self.setText('Произошла ошибка')
        self.setInformativeText(error_message)
        self.setStandardButtons(QMessageBox.StandardButton.Ok)


class QuestionMessageBox(QMessageBox):
    def __init__(self, question_text, parent=None):
        super().__init__(parent)
        self.setIcon(QMessageBox.Icon.Question)
        self.setText(question_text)
        self.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        self.setDefaultButton(QMessageBox.StandardButton.Yes)

    def exec_and_get_result(self):
        result = self.exec()
        return result == QMessageBox.StandardButton.Yes


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FileSelectionApp()
    window.show()
    sys.exit(app.exec())
